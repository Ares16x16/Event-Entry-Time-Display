import os
import openpyxl
import tkinter as tk
from tkinter import messagebox
import requests
import json

server = "http://localhost:8888/queue"

class ExcelEditor:
    def __init__(self, filename):
        self.filename = filename
        self.wb = openpyxl.load_workbook(filename)
        self.ws = self.wb.active

        self.current_value = self.ws.cell(row=1, column=1).value or 0
        self.current_timeslot = ""

    def increment_value(self):
        self.current_value += 1
        self.ws.cell(row=1, column=1).value = self.current_value
        self.wb.save(self.filename)
        self.value_label.config(text=str(self.current_value))

    def decrement_value(self):
        self.current_value -= 1
        self.ws.cell(row=1, column=1).value = self.current_value
        self.wb.save(self.filename)
        self.value_label.config(text=str(self.current_value))

    def update_timeslot(self):
        try:
            response = requests.get(server)
            response.raise_for_status()
            data = json.loads(response.content.decode())
            self.current_timeslot = data["current_timeslot"]
            self.timeslot_label.config(text=self.current_timeslot)
        except Exception as e:
            print(f"Error fetching timeslot: {e}")

    def create_gui(self):
        root = tk.Tk()
        root.title("Timeslot Editor")

        heading_label = tk.Label(root, text="Timeslot Editor", font=("Arial", 18), pady=10)
        heading_label.pack()

        value_frame = tk.Frame(root)
        value_label = tk.Label(value_frame, text=str(self.current_value), font=("Arial", 16), width=10)
        value_label.pack(side=tk.LEFT)
        self.value_label = value_label

        increment_button = tk.Button(value_frame, text="+", font=("Arial", 16), width=5, command=lambda: [self.increment_value(),self.update_timeslot()])
        increment_button.pack(side=tk.LEFT)

        decrement_button = tk.Button(value_frame, text="-", font=("Arial", 16), width=5, command=lambda: [self.decrement_value(),self.update_timeslot()])
        decrement_button.pack(side=tk.LEFT)

        value_frame.pack()

        timeslot_frame = tk.Frame(root)
        timeslot_label = tk.Label(timeslot_frame, text=self.current_timeslot, font=("Arial", 16), pady=10)
        timeslot_label.pack()
        self.timeslot_label = timeslot_label

        timeslot_button = tk.Button(timeslot_frame, text="Update Timeslot", font=("Arial", 14), command=self.update_timeslot)
        timeslot_button.pack()

        timeslot_frame.pack()

        root.mainloop()

if __name__ == "__main__":
    filename = "queue.xlsx"
    if not os.path.isfile(filename):
        messagebox.showerror("Error", "Excel file not found.")
        exit()

    editor = ExcelEditor(filename)
    editor.create_gui()
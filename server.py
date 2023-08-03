import socket
import threading
import logging
import json
import datetime
import openpyxl
import os
from datetime import datetime, timedelta, time

class VisitorManager:
    def __init__(self, filename, timeslot_duration=10, timeslot_capacity=50):
        self.timeslot_duration = timeslot_duration
        self.timeslot_capacity = timeslot_capacity
        self.current_number = None
        self.filename = filename
        self.wb = openpyxl.load_workbook(filename)
        self.ws = self.wb.active
        self.current_cell = self.ws.cell(row=1, column=1)
        self.last_modified = os.path.getmtime(filename)
        self.modified = False
        self.event_close_time = (18, 0) 

        self.current_number = int(self.current_cell.value)

    def get_current_timeslot(self):
        current_time = datetime.now().time()
        
        if self.last_modified < os.path.getmtime(self.filename):
            self.wb = openpyxl.load_workbook(self.filename) 
            self.ws = self.wb.active
            self.current_cell = self.ws.cell(row=1, column=1)
            self.last_modified = os.path.getmtime(self.filename) 
            self.current_number = int(self.current_cell.value)
            self.modified = True
            #print('last modified')
        
        if not self.modified and hasattr(self, 'cached_timeslot') and datetime.now() < self.cached_timeslot['expiry']:
            self.modified = False
            return self.cached_timeslot['timeslot']
        
        if datetime.combine(datetime.today(), current_time) >= datetime.combine(datetime.today(), time(hour=self.event_close_time[0], minute=self.event_close_time[1])):
            return "Venue is closed"
        
        current_number = self.get_current_number()
        timeslot_start = datetime.combine(datetime.today(), current_time) - timedelta(minutes=current_time.minute % self.timeslot_duration) + timedelta(minutes=(current_number - 1) * self.timeslot_duration) + timedelta(minutes=10)
        timeslot_end = timeslot_start + timedelta(minutes=self.timeslot_duration)
        timeslot = f"{timeslot_start.strftime('%H:%M')}-{timeslot_end.strftime('%H:%M')}"
        
        self.cached_timeslot = {'timeslot': timeslot, 'expiry': timeslot_end}
        
        return timeslot

    def get_current_timeslot_capacity(self):
        return self.timeslot_capacity

    def get_current_number(self):
        return int(self.current_cell.value)

    def to_dict(self):
        return {
            "current_timeslot": self.get_current_timeslot(),
            #"current_timeslot_capacity": self.get_current_timeslot_capacity(),
            #"current_number": self.get_current_number(),
        }


class WebServer:
    def __init__(self, host, port):
        self.host = host
        self.port = port
        self.webqueue = 0
        self.MAX_QUEUE_SIZE = 30
        self.queue = VisitorManager("queue.xlsx")
        self.lock = threading.Lock()
        self.logger = logging.getLogger("WebServer")
        self.logger.setLevel(logging.INFO)
        ch = logging.StreamHandler()
        # yyyy-mm-dd HH:MM:SS
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        ch.setFormatter(formatter)
        self.logger.addHandler(ch)
        
        self.index_data = self.load_file('index.html')
        self.css_data = self.load_file('style.css')
        self.js_data = self.load_file('app.js')
        self.queuePage_data = self.load_file('queuePage.html')
        
    def load_file(self, filename):
        with open(filename, 'r') as f:
            data = f.read()
        return data
    
    def start(self):
        self.socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        self.socket.bind((self.host, self.port))
        self.socket.listen(5)
        self.logger.info(f"Server started on {self.host}:{self.port}")
        threading.Thread(target=self.accept_connections, daemon=True).start()

    def accept_connections(self):
        while True:
            client_socket, client_address = self.socket.accept()
            self.logger.info(f"Accepted connection from {client_address}")
            threading.Thread(target=self.handle_connection, args=(client_socket,), daemon=True).start()

    def handle_connection(self, client_socket):
        
        self.webqueue += 1
        
        request = client_socket.recv(1024).decode()
        
        parsed_request = parse_http_request(request)
        method = parsed_request['method']
        path = parsed_request['path']
        self.logger.info(f"{method} {path}")

        if self.webqueue < self.MAX_QUEUE_SIZE:      
            if method == "GET" and path == "/":   
                response_data = self.index_data   
                response = f"HTTP/1.1 200 OK\r\n\r\n" + response_data
                self.logger.info(f"{method} {path} 200 OK")   
                
            elif method == "GET" and path == "/style.css":
                response_data = self.css_data   
                response = f"HTTP/1.1 200 OK\r\n\r\n" + response_data
                self.logger.info(f"{method} {path} 200 OK")  
                
            elif method == "GET" and path == "/app.js":
                response_data = self.js_data   
                response = f"HTTP/1.1 200 OK\r\n\r\n" + response_data
                self.logger.info(f"{method} {path} 200 OK")    
            
            elif method == "GET" and path == "/queue":
                response_data = self.queue.to_dict()
                response_body = json.dumps(response_data) 
                response = f"HTTP/1.1 200 OK\r\nContent-Type: application/json\r\n\r\n{response_body}"
                self.logger.info(f"{method} {path} 200 OK")  
                
            else:
                response = "HTTP/1.1 404 Not Found\r\n\r\n"
                self.logger.warning(f"404 Not Found")
        else:
            response_data = self.queuePage_data
            response = f"HTTP/1.1 200 OK\r\n\r\n" + response_data
            self.logger.info(f"Directed to queue page") 
            
        client_socket.sendall(response.encode())
        self.webqueue -= 1
        client_socket.close()
        
def parse_http_request(request_string):
    lines = request_string.split('\r\n')
    method, path, protocol = lines[0].split(' ', 2)
    headers = {}
    for line in lines[1:]:
        if not line:
            break
        key, value = line.split(': ', 1)
        headers[key] = value
    return {
        'method': method,
        'path': path,
        'protocol': protocol,
        'headers': headers
    }
    
if __name__ == "__main__":
    web_server = WebServer("0.0.0.0", 8888)
    web_server.start()

    while True:
        try:
            input()
        except KeyboardInterrupt:
            break

    web_server.stop()